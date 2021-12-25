import openpyxl, random
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.contrib.staticfiles.storage import staticfiles_storage
from .models import DailyMenu, Favorites

userList = ['김완수', '이영빈', '안태현', '전지훈', '박순재', '권용빈', '백종민', '홍석윤']

startCol = 8
userRowDict = dict(zip(userList, range(5, 20, 2)))

def exportWeeklyMenus(request):
    startDate = datetime.strptime(request.POST['startDate'], '%Y-%m-%d')
    endDate = startDate + timedelta(7)

    nthWeek = (startDate.replace(day = 1).weekday() + startDate.day) // 7 + 1

    response = HttpResponse(
        content_type='text/xlsx',
        headers={'Content-Disposition': 'attachment; filename="meal.xlsx"'},
    )

    formFilePath = staticfiles_storage.path('meal/sheet/form.xlsx')

    wb = openpyxl.load_workbook(formFilePath)
    ws1 = wb['개인']
    ws2 = wb['주문']

    ws1.cell(2, startCol, f"{startDate.year}년 {startDate.month}월 {nthWeek}주차 조식({startDate.strftime('%m/%d')}~{endDate.strftime('%m/%d')})")
    for i in range(7):
        date = startDate + timedelta(i)
        ws1.cell(4, 8 + 3 * i, date.strftime('%m/%d'))
        ws2.cell(2, 5 + 5 * i, date.strftime('%m/%d'))
        menuList = DailyMenu.objects.filter(date = date.strftime('%Y-%m-%d'))
        for obj in menuList:
            row, col = userRowDict[str(obj.user)], startCol + 3 * i
            favoritesList = Favorites.objects.filter(user = obj.user)
            if len(favoritesList) == 0:
                favoritesList = Favorites.objects.all()

            if obj.status >= 1:
                menus = list(filter(lambda a: a != None, [obj.menu1, obj.menu2, obj.menu3]))
            else:
                randomFavorites = favoritesList[random.randrange(len(favoritesList))]
                menus = list(filter(lambda a: a != None, [randomFavorites.menu1, randomFavorites.menu2, randomFavorites.menu3]))
                obj.menu1 = menus[0]
                obj.menu2 = menus[1]
                if len(menus) == 3:
                    obj.menu3 = menus[2]
                obj.status = 1
                obj.save()
                
            if len(menus) == 0:
                ws1.cell(row, col, '결식')
                ws1.cell(row + 1, col, 0)
                ws1.merge_cells(start_row = row, start_column = col, end_row = row, end_column = col + 2)
                ws1.merge_cells(start_row = row + 1, start_column = col, end_row = row + 1, end_column = col + 2)
            elif len(menus) == 1:
                ws1.cell(row, col, menus[0])
                ws1.merge_cells(start_row = row, start_column = col, end_row = row, end_column = col + 2)
                ws1.merge_cells(start_row = row + 1, start_column = col, end_row = row + 1, end_column = col + 2)
            elif len(menus) == 2:
                ws1.cell(row, col, menus[0])
                ws1.cell(row, col + 2, menus[1])
                ws1.merge_cells(start_row = row, start_column = col, end_row = row, end_column = col + 1)
                ws1.merge_cells(start_row = row + 1, start_column = col, end_row = row + 1, end_column = col + 1)
            else:
                for j, menu in enumerate(menus):
                    ws1.cell(row, col + j, menu)

    wb.save(response)
  
    return response